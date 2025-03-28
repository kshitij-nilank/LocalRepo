import pandas as pd
from pandas.api.types import CategoricalDtype
import datetime
import numpy as np
import mysql.connector as msql
from mysql.connector import Error
from google.cloud import bigquery
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import  Alignment, Border, Side, Font, PatternFill
from openpyxl.cell.cell import MergedCell

import os 
os.chdir(r"D:\Spyder_Codes")
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'D:\GCP_Auth\data-warehousing-prod-bc4ee5babbb0.json'

#Imports google cloud client library and initiates BQ service
from google.cloud import bigquery
#from google.cloud import datastore
bigquery_client = bigquery.Client()

QUERY = """
SELECT
    Centre, FinYear,Season, SaleNo, AuctionDate, LotNo, 
Garden, GardenMDM,Grade, GradeMDM, InvoiceNo, 
Buyer, BuyerMDM, BuyerGroup, BrokerCode,
Seller, SellerGroup, Category, SubCategory, TeaType, 
SubTeaType,LotStatus, Area, EstBlf,GPDATE,ReprintNo,
SUM(IF(LotStatus = 'Sold',TotalWeight,InvoiceWeight)) AS Offer_Qty,
SUM(TotalWeight) AS Sold_Qty,
SUM(Value) AS Total_Value

FROM `data-warehousing-prod.EasyReports.SaleTransactionView`

WHERE Season IN (2024) And SellerGroup = "LUXMI"

GROUP BY Centre, FinYear,Season, SaleNo, AuctionDate, LotNo, Garden, GardenMDM,Grade, GradeMDM, 
InvoiceNo, Buyer, BuyerMDM, BuyerGroup, BrokerCode,Seller, SellerGroup, Category, SubCategory, 
TeaType, SubTeaType, LotStatus, Area, EstBlf,GPDATE,ReprintNo """

Query_Results = bigquery_client.query(QUERY)
df = Query_Results.to_dataframe()

df['Avg_Price'] = np.where(df['Sold_Qty'] > 0, df['Total_Value'] / df['Sold_Qty'], 0)
#df['Out%'] = (1-(df['Sold_Qty'] / df['Offer_Qty']))*100
df['SaleAlies'] = np.where((df['SaleNo'] >= 1) & (df['SaleNo'] <= 13), df['SaleNo'] + 52, df['SaleNo'])

df.info()

df1=df[(df['Category'].isin(["CTC"])) & (df['EstBlf']=="EST") & (df['SaleAlies'].between(14,60))]

# Step 1: Aggregate Data Before Pivoting
summary_df = (df1.groupby(["SubTeaType", "GradeMDM", "GardenMDM"]).agg({
    "Offer_Qty":"sum","Sold_Qty":"sum","Avg_Price":"mean"}).reset_index())

##########
# Define primary and secondary grade orders explicitly
primary_grades = ["BOPL", "BPS", "BOP", "BOPSM", "BPSM", "BP", "PF", "OF", "PD", "D", "CD"]
secondary_grades = ["BOPL1", "BPS1", "BOP1", "BOPSM1", "BP1", "PF1", "OF1", "PD1", "D1", "CD1"]

# Create a mapping dictionary for grade sorting
grade_order_dict = {grade: i for i, grade in enumerate(primary_grades + secondary_grades)}

# Assign category type: 1 for Primary, 2 for Secondary
summary_df["Category_Order"] = summary_df["GradeMDM"].apply(lambda x: 1 if x in primary_grades else 2)

# Assign sorting order based on predefined grade order
summary_df["Grade_Order"] = summary_df["GradeMDM"].map(grade_order_dict)

# Sort first by Category (Primary → Secondary), then by Grade order
summary_df = summary_df.sort_values(by=["Category_Order", "Grade_Order"])

# Drop the helper columns after sorting
summary_df = summary_df.drop(columns=["Category_Order", "Grade_Order"])




##########

# Step 2: Create Pivot Table with Multi-Index
pivot_df = summary_df.pivot_table(
    index=["SubTeaType", "GradeMDM"],  # Multi-index (SubTeaType -> GradeMDM)
    columns="GardenMDM",  # Columns
    values=["Offer_Qty","Sold_Qty", "Avg_Price"],  # Metrics to show
    fill_value=0,  # Replace NaNs with 0
    aggfunc={"Offer_Qty":"sum","Sold_Qty": "sum", "Avg_Price": "mean"}).round(0)

#SWAPING
pivot_df=pivot_df.swaplevel(axis=1).sort_index(axis=1)

#################Creating Percentage#################
sold_qty = pivot_df.xs('Sold_Qty', axis=1, level=1)

# Calculate the parent total for each GardenMDM
total_sold_qty = sold_qty.sum()

# Calculate the percentage of parent total
pct_of_parent = sold_qty.divide(total_sold_qty, axis=1) * 100

# Add the percentage of parent total to the original DataFrame
for garden in pct_of_parent.columns:
    pivot_df[(garden, 'Grade%')] = pct_of_parent[garden]

# Sort columns for better readability
pivot_df = pivot_df.sort_index(axis=1)

#checking the performance of percentage
pivot_df.loc[:, pivot_df.columns.get_level_values(1) == 'Grade%'].sum()

#################Creating Out%#################

offer_qty = pivot_df.xs('Offer_Qty', axis=1, level=1)

# Calculate 'Out%' using the given formula
out_percentage = (1 - (sold_qty / offer_qty)) * 100

# Handle division by zero (if Offer_Qty is 0)
out_percentage = out_percentage.fillna(0)

# Add 'Out%' back to pivot_df
for garden in out_percentage.columns:
    pivot_df[(garden, 'Out%')] = out_percentage[garden]

# Sort columns for better readability
pivot_df = pivot_df.sort_index(axis=1)

# Checking the performance of 'Out%'
pivot_df.loc[:, pivot_df.columns.get_level_values(1) == 'Out%'].mean()

###########################
sum_cols = ['Sold_Qty', 'Grade%']
avg_cols = ['Avg_Price']
weight_col = 'Sold_Qty'

def add_subtotals(df):
    subtotals = []
    
    for category in ['PRIMARY', 'SECONDARY']:
        category_rows = df.loc[category]
        subtotal = pd.DataFrame(index=[(category, 'SubTotal')], columns=df.columns)
        
        for garden in df.columns.levels[0]:  
            # Sum Sold_Qty & Grade%
            subtotal[(garden, 'Sold_Qty')] = category_rows[(garden, 'Sold_Qty')].sum()
            subtotal[(garden, 'Grade%')] = category_rows[(garden, 'Grade%')].sum()
            
            # Weighted Average for Avg_Price
            total_weight = category_rows[(garden, weight_col)].sum()
            if total_weight > 0:
                subtotal[(garden, 'Avg_Price')] = (
                    (category_rows[(garden, 'Avg_Price')] * category_rows[(garden, weight_col)]).sum() / total_weight
                )
            else:
                subtotal[(garden, 'Avg_Price')] = 0

            # Sum Offer_Qty (Needed for Out% Calculation)
            subtotal[(garden, 'Offer_Qty')] = category_rows[(garden, 'Offer_Qty')].sum()

            # Correct Out% Calculation
            if subtotal[(garden, 'Offer_Qty')].iloc[0] > 0:
                subtotal[(garden, 'Out%')] = (1 - (subtotal[(garden, 'Sold_Qty')] / subtotal[(garden, 'Offer_Qty')])) * 100
            else:
                subtotal[(garden, 'Out%')] = 0  # Avoid division by zero
        
        subtotals.append(subtotal)

    df = pd.concat([df] + subtotals).sort_index()

    # Compute Grand Total
    grand_total = pd.DataFrame(index=[('Grand Total', '')], columns=df.columns)

    for garden in df.columns.levels[0]:  
        grand_total[(garden, 'Sold_Qty')] = df.xs('SubTotal', level=1).loc[:, (garden, 'Sold_Qty')].sum()
        grand_total[(garden, 'Grade%')] = df.xs('SubTotal', level=1).loc[:, (garden, 'Grade%')].sum()
        grand_total[(garden, 'Offer_Qty')] = df.xs('SubTotal', level=1).loc[:, (garden, 'Offer_Qty')].sum()

        total_weight = df.xs('SubTotal', level=1).loc[:, (garden, weight_col)].sum()
        if total_weight > 0:
            grand_total[(garden, 'Avg_Price')] = (
                (df.xs('SubTotal', level=1).loc[:, (garden, 'Avg_Price')] * df.xs('SubTotal', level=1).loc[:, (garden, weight_col)]).sum()
                / total_weight
            )
        else:
            grand_total[(garden, 'Avg_Price')] = 0
        
        # Correct Out% Calculation for Grand Total
        if grand_total[(garden, 'Offer_Qty')].iloc[0] > 0:
            grand_total[(garden, 'Out%')] = (1 - (grand_total[(garden, 'Sold_Qty')] / grand_total[(garden, 'Offer_Qty')])) * 100
        else:
            grand_total[(garden, 'Out%')] = 0

    df = pd.concat([df, grand_total])
    
    return df

# Apply function
pivot_df = add_subtotals(pivot_df)

###########################

######### Remove Offer_Qty#########
pivot_df = pivot_df.drop(columns=pivot_df.columns[pivot_df.columns.get_level_values(1) == 'Offer_Qty'])

################################

desired_order = ['Sold_Qty', 'Avg_Price', 'Grade%', 'Out%']

# Rearrange columns under each Garden name
new_columns = []
for garden in pivot_df.columns.levels[0]:  # Iterate over the garden names
    for metric in desired_order:  # Maintain the desired metric order
        new_columns.append((garden, metric))

# Update the DataFrame with the new column order
pivot_df = pivot_df[new_columns]


with pd.ExcelWriter("final_result.xlsx", engine="openpyxl") as writer:
    pivot_df.to_excel(writer, sheet_name="Summary")

    # Load the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets["Summary"]

    # Define border style (thin border for all sides)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Define bold font for subtotal and grand total
    bold_font = Font(bold=True)

    # Define light green fill for subotal and grand total rows
    light_green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

      # Loop through all rows and columns
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
        first_cell = row[0]  # First cell of the row (index column)

        # Check if row is "SubTotal" or "Grand Total"
        if first_cell.value and ("SubTotal" in str(first_cell.value) or "Grand Total" in str(first_cell.value)):
            for cell in row:  # Apply to all columns in that row
                cell.font = bold_font
                cell.fill = light_green_fill

          
    # Save the workbook
    workbook.save("final_result.xlsx")
    
    #addition of new commit